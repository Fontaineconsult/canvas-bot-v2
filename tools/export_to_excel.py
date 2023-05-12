import os
import zipfile
from collections import namedtuple
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


tracking_columns = {
    'Documents': [
        ('Requires OCR', 12, "No"),
        ('Accessible', 15, 'Not Checked'),
        ('Accessible File Location', 20, None),
        ('Notes', 20, None),
        ('Ignore Doc', 20, "No"),


    ],
    'Document Sites': [
        ('ColumnA', 12, None),
        ('ColumnB', 18, None),
        ('Notes', 20, None),
        ('Ignore', 20, "No"),
    ],
    'Image Files': [
        ('Needs Description?', 12, None),
        ('Notes', 20, None),
        ('Ignore', 20, "No"),
    ],
    'Video Files': [
        # ('Captioned', 15),
        ('Accessible File Location', 20, None),
        ('Sent To AST', 20, "No"),
        ('Notes', 20, None),
        ('Ignore', 20, "No"),
    ],
    'Video Sites': [
        # ('Captioned', 15),
        ('Accessible Video Location', 20, None),
        ('Sent To AST', 20, "No"),
        ('Notes', 20, None),
        ('Ignore Video', 20, "No"),
    ],
    'Audio Files': [
        ('Captioned', 15, None),
        ('Accessible File Location', 20, None),
        ('Sent To AST', 20, None),
        ('Notes', 20, None),
        ('Ignore', 20, "No"),
    ],
    'Audio Sites': [
        ('Transcript', 15, None),
        ('Accessible Audio Location', 20, None),
        ('Sent To AST', 20, None),
        ('Notes', 20, None),
        ('Ignore', 20, "No"),
    ],
    'Unsorted': [
    ],
}


data_validations = {
    # keys are final column names
    'Accessible': [
        (DataValidation(
            type='list',
            formula1='"{}"'.format(','.join(['Not Checked',
                                             'Not Accessible',
                                             'Already Accessible',
                                             'Remediated to DPRC Minimum (PDF UA Partially Compliant)',
                                             'Fully Remediated To WCAG 2.0 AA',
                                             'Unable to Remediate'])),
            showDropDown=False),
         [
             FormulaRule(formula=['$K2="Not Checked"'],
                        fill=PatternFill(bgColor='FFFF00', fgColor='FFFF00', fill_type='solid')),
             FormulaRule(formula=['$K2="Remediated to DPRC Minimum (PDF UA Partially Compliant)"'],
                        fill=PatternFill(bgColor='FFA143', fgColor='FFFF00', fill_type='solid')),
             FormulaRule(formula=['$K2="Remediated To WCAG 2.0 AA"'],
                        fill=PatternFill(bgColor='92D050', fgColor='FFFF00', fill_type='solid')),
             FormulaRule(formula=['$K2="Unable to Remediate"'],
                         fill=PatternFill(bgColor='DB3535', fgColor='FFFF00', fill_type='solid')),
             FormulaRule(formula=['$K2="Not Accessible"'],
                         fill=PatternFill(bgColor='DB3535', fgColor='FFFF00', fill_type='solid')),
             FormulaRule(formula=['$K2="Already Accessible"'],
                         fill=PatternFill(bgColor='92D050', fgColor='FFFF00', fill_type='solid')),
         ]
        )
    ],
    'Caption Status': [
        (
            DataValidation(
                type='list',
                formula1='"{}"'.format(','.join(['Not Checked',
                                                 'Captioned',
                                                 'Auto Caption',
                                                 'Not Captioned'])),
                showDropDown=False),
         [
             FormulaRule(formula=['$A2="Not Checked"'],
                         fill=PatternFill(bgColor='FFFF00', fgColor='FFFF00', fill_type='solid')),
             FormulaRule(formula=['$A2="Captioned"'],
                         fill=PatternFill(bgColor='92D050', fgColor='FFFF00', fill_type='solid')),
             FormulaRule(formula=['$A2="Auto Caption"'],
                         fill=PatternFill(bgColor='FFA143', fgColor='FFA143', fill_type='solid')),
             FormulaRule(formula=['$A2="Not Captioned"'],
                         fill=PatternFill(bgColor='DB3535', fgColor='DB3535', fill_type='solid')),

         ]
        )
    ],
    'Sent To AST': [(DataValidation(
        type='list',
        formula1='"{}"'.format(','.join(['Yes',
                                         'No'])),
        showDropDown=False),
                     [
                         FormulaRule(formula=['$J2="Yes"'],
                                     fill=PatternFill(bgColor='FFFF00', fgColor='FFFF00', fill_type='solid')),
                         FormulaRule(formula=['$J2="No"'],
                                     fill=PatternFill(bgColor='92D050', fgColor='FFFF00', fill_type='solid')),
                     ])],
    'Ignore Doc': [(DataValidation(
        type='list',
        formula1='"{}"'.format(','.join(['Yes',
                                         'No'])),
        showDropDown=False),
                     [
                         FormulaRule(formula=['$N2="Yes"'],
                                     fill=PatternFill(bgColor='FFFF00', fgColor='FFFF00', fill_type='solid')),
                         FormulaRule(formula=['$N2="No"'],
                                     fill=PatternFill(bgColor='92D050', fgColor='FFFF00', fill_type='solid')),
                     ])],
    'Ignore Video': [(DataValidation(
        type='list',
        formula1='"{}"'.format(','.join(['Yes',
                                         'No'])),
        showDropDown=False),
                    [
                        FormulaRule(formula=['$L2="Yes"'],
                                    fill=PatternFill(bgColor='FFFF00', fgColor='FFFF00', fill_type='solid')),
                        FormulaRule(formula=['$L2="No"'],
                                    fill=PatternFill(bgColor='92D050', fgColor='FFFF00', fill_type='solid')),
                    ])],
    'Requires OCR': [(DataValidation(
        type='list',
        formula1='"{}"'.format(','.join(['Yes',
                                         'No'])),
        showDropDown=False),
                [
                    FormulaRule(formula=['$J2="Yes"'],
                                fill=PatternFill(bgColor='FFFF00', fgColor='FFFF00', fill_type='solid')),
                    FormulaRule(formula=['$J2="No"'],
                                fill=PatternFill(bgColor='92D050', fgColor='FFFF00', fill_type='solid')),
                ])]


}


pattern_fills = {
    # key is from data dict
    'is_hidden': [(False, PatternFill(bgColor='92D050', fgColor='92D050', fill_type='solid'), "Visible"),
                  (True, PatternFill(bgColor='DB3535', fgColor='DB3535', fill_type='solid'), "Hidden"),
                  ],


}



def add_validation_formatting(sheet, validations, col_idx):



    for v in validations:
        validation = v[0]
        formatting_list = v[1]

        cell_range = get_data_cells_range(sheet, 5)

        validation_cell_range = replace_column_in_range(cell_range, get_column_letter(col_idx))

        sheet.add_data_validation(validation)
        validation.ranges.add(validation_cell_range)

        for formatting in formatting_list:
            formatter = formatting
            sheet.conditional_formatting.add(validation_cell_range, formatter)



def replace_column_in_range(range_string, new_column_letter):
    # Split the range string into its parts
    start, end = range_string.split(':')

    # Extract the row numbers
    start_row = ''.join(filter(str.isdigit, start))
    end_row = ''.join(filter(str.isdigit, end))

    # Replace the column letter in the range string
    new_range_string = f"{new_column_letter}{start_row}:{new_column_letter}{end_row}"
    return new_range_string


def has_data(sheet):
    # Check if there's data in any cell of the sheet
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                return True
    return False


def get_data_cells_range(sheet, column):
    start_row = 2
    end_row = find_next_empty_row(sheet, column) - 1

    if end_row < start_row:
        return None  # No data cells in the column

    start_cell = sheet.cell(row=start_row, column=column)
    end_cell = sheet.cell(row=end_row, column=column)

    return f"{start_cell.coordinate}:{end_cell.coordinate}"


def find_next_empty_column(sheet):
    col_idx = 1
    while sheet.cell(row=1, column=col_idx).value is not None:
        col_idx += 1
    return col_idx

def find_next_empty_row(sheet, column):
    row_idx = 1
    while sheet.cell(row=row_idx, column=column).value is not None:
        row_idx += 1
    return row_idx


def get_sheet_column_names(sheet):
    print(sheet)
    print(sheet[1])
    return [cell.value for cell in sheet[1]]


def create_excel_file(json_data, excel_file_path=None):

    """
    Creates an excel file with the appropriate sheets.
    :param json_data:
    :param excel_file_path:
    :return:
    """
    wb = openpyxl.Workbook()
    ws = wb["Sheet"]
    wb.remove(ws)
    wb.create_sheet('Documents')
    wb.create_sheet('Document Sites')
    wb.create_sheet('Image Files')
    wb.create_sheet('Video Files')
    wb.create_sheet('Video Sites')
    wb.create_sheet('Audio Files')
    wb.create_sheet('Audio Sites')
    wb.create_sheet('Unsorted')

    wb.save(excel_file_path)
    wb.close()

    # build_xcel_file(json_data, excel_file_path)


def remove_key_recursively(obj, key_to_remove):
    if isinstance(obj, dict):
        result = {}
        for key, value in obj.items():
            if key != key_to_remove:
                result[key] = remove_key_recursively(value, key_to_remove)
        return result
    elif isinstance(obj, list):
        result = []
        for item in obj:
            result.append(remove_key_recursively(item, key_to_remove))
        return result
    else:
        return obj


def apply_sheet_styles(excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)



    for sheet_name in wb.sheetnames:

        sheet = wb[sheet_name]

        if sheet.dimensions == "A1:A1":
            continue

        # first_row_values = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))][0]
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 30
        sheet.column_dimensions['H'].width = 40
        sheet_name = sheet_name.replace(" ", "-")


        table = Table(displayName=sheet_name, ref=sheet.dimensions)

        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)

        table.tableStyleInfo = style
        sheet.add_table(table)

    wb.save(excel_file_path)
    wb.close()


def find_key_names(d, path=None):
    if path is None:
        path = []
    result = []
    for k, v in d.items():
        new_path = path + [k]
        if isinstance(v, list):
            result.append(new_path)
        elif isinstance(v, dict):
            sub_result = find_key_names(v, new_path)
            result += sub_result
    return result


def add_header_to_sheet(file_path, sheet_name, header_row):
    """
    Adds a header row to a sheet based on the keys in a Python dictionary.

    Args:
        file_path (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to add the header row to.
        header_row (dict): A dictionary whose keys will be used as the header row.
    """
    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # Add the header row to the sheet

    for col_idx, header_name in enumerate(header_row):

        cell = sheet.cell(row=1, column=col_idx+1)
        cell.value = " ".join(word.capitalize() for word in header_name.replace('_', ' ').split(' '))

    # Save the workbook
    wb.save(file_path)
    wb.close()


def dicts_to_excel(filename, sheetname, data, download_hidden_files):
    # Load an existing Excel workbook or create a new one if it doesn't exist
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()


    # Create a new sheet with the specified sheet name or get the existing one
    if sheetname in wb:
        ws = wb[sheetname]
    else:
        ws = wb.create_sheet(sheetname)

    # Write data rows (values from each dictionary)

    for row_num, item in enumerate(data, 2):  # Start from row 2

        for col_num, key in enumerate(item.keys(), 1):

            if key == "save_path":
                if item['is_hidden'] and not download_hidden_files:
                    continue

                target = item.get(key)
                if item['source_page_type'] == 'BoxPage':
                    target = item.get("url")


                hyperlink = Hyperlink(ref=f"{get_column_letter(col_num)}{row_num}",
                                      target=target, display="Open File")
                ws.cell(row=row_num, column=col_num).hyperlink = hyperlink
                continue

            # pattern fills
            pattern_fill = pattern_fills.get(key)
            if pattern_fill:
                for pattern in pattern_fill:
                    if item.get(key) == pattern[0]:
                        fill = pattern[1]
                        ws.cell(row=row_num, column=col_num).fill = fill
                        ws.cell(row=row_num, column=col_num).value = pattern[2]
                continue

            ws.cell(row=row_num, column=col_num).value = item.get(key)


    # Save the workbook to a file
    wb.save(filename)
    wb.close()


def build_xcel_file(json_data, excel_file_path, download_hidden_files):

    key_paths = find_key_names(json_data)
    for path in key_paths:
        # navigate to the list using each path
        sub_dict = json_data
        for key in path[:-1]:
            sub_dict = sub_dict[key]

        my_list = sub_dict[path[-1]]
        sheet_name = " ".join(word.capitalize() for word in path[-1].replace('_', ' ').split(' '))

        try:
            add_header_to_sheet(excel_file_path, sheet_name, my_list[0])
            dicts_to_excel(excel_file_path, sheet_name, my_list, download_hidden_files)
        except IndexError:
            pass

def add_tracking_columns(excel_file_path):


    wb = openpyxl.load_workbook(excel_file_path)


    for sheet_name, columns in tracking_columns.items():
        sheet = wb[sheet_name]

        # Only update the sheet if it has data
        if has_data(sheet):
            start_col_idx = find_next_empty_column(sheet)
            for col_offset, (col_title, col_width, default_value) in enumerate(columns):
                col_idx = start_col_idx + col_offset

                # Set column title
                sheet.cell(row=1, column=col_idx, value=col_title)

                # Set data validations
                # validations = data_validations.get(col_title)
                # if data_validations.get(col_title):
                #     add_validation_formatting(sheet, data_validations.get(col_title), col_idx)


                # if validations is not None:
                #     for v in validations:
                #         validation = v[0]
                #         formatting_list = v[1]
                #
                #         cell_range = get_data_cells_range(sheet, 5)
                #
                #         validation_cell_range = replace_column_in_range(cell_range, get_column_letter(col_idx))
                #
                #         sheet.add_data_validation(validation)
                #         validation.ranges.add(validation_cell_range)
                #
                #         for formatting in formatting_list:
                #             formatter = formatting
                #             sheet.conditional_formatting.add(validation_cell_range, formatter)
                #
                #         for row in sheet[validation_cell_range]:
                #             for cell in row:
                #                 print(cell)
                #                 cell.value = "Not Checked"


                # Set column width (optional)
                if col_width is not None:
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    sheet.column_dimensions[column_letter].width = col_width

                if default_value is not None:
                    cell_range = get_data_cells_range(sheet, 5)
                    tracking_cell_range = replace_column_in_range(cell_range, get_column_letter(col_idx))
                    for row in sheet[tracking_cell_range]:
                        for cell in row:
                            cell.value = default_value

    wb.save(excel_file_path)
    wb.close()



def add_data_validations(excel_file_path):

    wb = openpyxl.load_workbook(excel_file_path)

    for sheet in wb.sheetnames:
        active_sheet = wb[sheet]
        col_headers = get_sheet_column_names(active_sheet)
        if len(col_headers) > 0:
            for col_idx, col_title in enumerate(col_headers, 1):
                print(col_idx, col_title)

                if data_validations.get(col_title):
                    validation = data_validations.get(col_title)

                    for v in validation:
                        validation = v[0]
                        formatting_list = v[1]

                        cell_range = get_data_cells_range(active_sheet, 5)

                        validation_cell_range = replace_column_in_range(cell_range, get_column_letter(col_idx))

                        active_sheet.add_data_validation(validation)
                        validation.ranges.add(validation_cell_range)

                        for formatting in formatting_list:
                            formatter = formatting
                            active_sheet.conditional_formatting.add(validation_cell_range, formatter)





    wb.save(excel_file_path)
    wb.close()


def save_as_excel(json_data, file_save_path, download_hidden_files):

    xcel_path = os.path.join(file_save_path, json_data['course_id'] + '.xlsxm')
    json_data = remove_key_recursively(json_data, 'path')

    create_excel_file(json_data, xcel_path)
    build_xcel_file(json_data, xcel_path, download_hidden_files)
    add_tracking_columns(xcel_path)
    add_data_validations(xcel_path)
    apply_sheet_styles(xcel_path)

